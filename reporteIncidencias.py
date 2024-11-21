from openpyxl import Workbook
from openpyxl.styles import Border,Side
from openpyxl.styles.alignment import Alignment
from xml.etree.ElementTree import Element, SubElement, tostring
from et_xmlfile import xmlfile
import Consulta_Triaje,Consulta_Galen
from tkinter import messagebox
from tkinter import filedialog
from datetime import datetime

class Reporte(object):
	def __init__(self):
		double_border_side=Side(border_style="thin")
		self.borde_caja=Border(top=double_border_side,right=double_border_side,bottom=double_border_side,left=double_border_side)
		self.borde_superior=Border(top=Side(border_style='thick'))
		self.obj_consulta=Consulta_Triaje.queryTriaje()
		self.obj_consultaG=Consulta_Galen.queryGalen()
	def Genera_RDatos(self,fechaI,fechaF,indicador):
		
		ruta=filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Archivos de Texto", "*.xlsx"), ("Todos los archivos", "*.*")])
		if len(ruta)>0:
			rows=self.obj_consulta.ReporteNoAtendidos(fechaI,fechaF)
			if indicador=='FILTRADO':
				self.Report_DigitD(fechaI,fechaF,rows,ruta)
			elif indicador=='GENERAL':
				self.Report_General(fechaI,fechaF,rows,ruta)
		else:
			messagebox.showerror("Alerta","Ubicacion no válida")

	def Report_DigitD(self,fechaI,fechaF,rows,ruta):
		
		wb=Workbook()
		sheet=wb.active
		#Encabezado de la hoja HIS
		sheet.merge_cells('A1:E1')

		sheet.column_dimensions['A'].width=10
		sheet.column_dimensions['B'].width=50
		sheet.column_dimensions['C'].width=30
		sheet.column_dimensions['D'].width=20
		sheet.column_dimensions['E'].width=15

		#sheet['A1'].border=self.borde_caja
		sheet['A1']="REPORTE DE PACIENTES NO ATENDIDOS"
		sheet['A1'].alignment=Alignment(horizontal="center")

		sheet['A2']="DNI"
		sheet['A2'].alignment=Alignment(horizontal="center")

		sheet['B2']="NOMBRES Y APELLIDOS"
		sheet['B2'].alignment=Alignment(horizontal="center")

		sheet['C2']="ESPECIALIDAD"
		sheet['C2'].alignment=Alignment(horizontal="center")

		sheet['D2']="TELEFONO"
		sheet['D2'].alignment=Alignment(horizontal="center")

		sheet['E2']="FECHA REGISTRO"
		sheet['E2'].alignment=Alignment(horizontal="center")

		sheet['F2']="MOTIVO"
		sheet['F2'].alignment=Alignment(horizontal="center")

		nro=3
		for val in rows:
			rowsGalen=self.obj_consultaG.query_NoAtendidos(fechaI,fechaF,val.Especialidad,val.Dni_Paciente)		
			
			if len(rowsGalen)==0:				
				dt=self.obj_consultaG.query_PacienteSindireccion(val.Dni_Paciente)
				TriajePaciente=self.obj_consulta.Consulta_DNIPaciente(val.Dni_Paciente)

				if not TriajePaciente==None:
					sheet['B'+str(nro)]=TriajePaciente.Nombre+" "+TriajePaciente.Apellido_Paterno+" "+TriajePaciente.Apellido_Materno
				else:
					for v in dt:					
						sheet['D'+str(nro)]=v.Telefono
						sheet['B'+str(nro)]=v.PrimerNombre+" "+v.ApellidoPaterno+" "+v.ApellidoMaterno

				#sheet['B'+str(nro)]=""					
				sheet['A'+str(nro)]=val.Dni_Paciente				
				sheet['C'+str(nro)]=val.Especialidad				
				sheet['E'+str(nro)]=val.Fecha
				sheet['F'+str(nro)]=val.Motivo
				nro=nro+1			
					
		wb.save(f"{ruta}")

	def Report_General(self,fechaI,fechaF,rows,ruta):
		
		wb=Workbook()
		sheet=wb.active
		#Encabezado de la hoja HIS
		sheet.merge_cells('A1:E1')

		sheet.column_dimensions['A'].width=10
		sheet.column_dimensions['B'].width=50
		sheet.column_dimensions['C'].width=30
		sheet.column_dimensions['D'].width=20
		sheet.column_dimensions['E'].width=15

		#sheet['A1'].border=self.borde_caja
		sheet['A1']="REPORTE DE PACIENTES NO ATENDIDOS"
		sheet['A1'].alignment=Alignment(horizontal="center")

		sheet['A2']="DNI"
		sheet['A2'].alignment=Alignment(horizontal="center")

		sheet['B2']="NOMBRES Y APELLIDOS"
		sheet['B2'].alignment=Alignment(horizontal="center")

		sheet['C2']="ESPECIALIDAD"
		sheet['C2'].alignment=Alignment(horizontal="center")

		sheet['D2']="TELEFONO"
		sheet['D2'].alignment=Alignment(horizontal="center")

		sheet['E2']="FECHA REGISTRO"
		sheet['E2'].alignment=Alignment(horizontal="center")

		sheet['F2']="MOTIVO"
		sheet['F2'].alignment=Alignment(horizontal="center")

		nro=3
		for val in rows:
			dt=self.obj_consultaG.query_PacienteSindireccion(val.Dni_Paciente)
			TriajePaciente=self.obj_consulta.Consulta_DNIPaciente(val.Dni_Paciente)

			if not TriajePaciente==None:
				sheet['B'+str(nro)]=TriajePaciente.Nombre+" "+TriajePaciente.Apellido_Paterno+" "+TriajePaciente.Apellido_Materno
			else:
				for v in dt:
					sheet['D'+str(nro)]=v.Telefono
					sheet['B'+str(nro)]=v.PrimerNombre+" "+v.ApellidoPaterno+" "+v.ApellidoMaterno


			#sheet['B'+str(nro)]=""					
			sheet['A'+str(nro)]=val.Dni_Paciente				
			sheet['C'+str(nro)]=val.Especialidad				
			sheet['E'+str(nro)]=val.Fecha
			sheet['F'+str(nro)]=val.Motivo
			nro=nro+1			
					
		wb.save(f"{ruta}")