from tkcalendar import DateEntry
from tkinter import *
from tkinter import messagebox
from tkinter import ttk,filedialog
from openpyxl import Workbook
from openpyxl.styles import Border,Side
from openpyxl.styles.alignment import Alignment
from Nacidos.consultaN import Consulta
from Consulta_Galen import queryGalen
class Reporte():
	def __init__(self):
		double_border_side=Side(border_style="thin")
		self.borde_caja=Border(top=double_border_side,right=double_border_side,bottom=double_border_side,left=double_border_side)
	def TopReporte(self):
		self.Top_Reporte=Toplevel()
		self.Top_Reporte.geometry("400x150")
		self.Top_Reporte.title("Reporte AIRN")
		self.Top_Reporte.resizable(0,0)
		self.Top_Reporte.grab_set()

		label=Label(self.Top_Reporte,text="Desde: ",font=('Arial',10,'bold'))
		label.grid(row=1,column=1,pady=10)	
		self.fechaDesde=DateEntry(self.Top_Reporte,selectmode='day',date_pattern='yyyy-MM-dd')
		self.fechaDesde.grid(row=1,column=2,pady=10,padx=10)

		label=Label(self.Top_Reporte,text="Desde: ",font=('Arial',10,'bold'))
		label.grid(row=1,column=3,pady=10)	
		self.fechaHasta=DateEntry(self.Top_Reporte,selectmode='day',date_pattern='yyyy-MM-dd')
		self.fechaHasta.grid(row=1,column=4,pady=10,padx=10)

		buttonGENERAR=ttk.Button(self.Top_Reporte,text="Generar")
		buttonGENERAR.grid(row=2,column=3,pady=10)
		buttonGENERAR['command']=self.evento_Button

	def evento_Button(self):
		file_Address=filedialog.asksaveasfile(mode="w",defaultextension=".xlsx")
		self.ReporteNacidos(file_Address)
		self.Top_Reporte.destroy()
		messagebox.showinfo("Notificación","Se generó correctamente!!")

	def ReporteNacidos(self,address):
		obj_consulta=Consulta()
		obj_ConsultaGalen=queryGalen()
		fechaI=self.fechaDesde.get_date()
		fechaF=self.fechaHasta.get_date()

		rows=obj_consulta.Report_GeneralRN(fechaI,fechaF)
		wb=Workbook()
		sheet=wb.active
		sheet.merge_cells('A1:T1')		
		sheet['A1']=f"REGISTRO DE SEGUIMIENTO DE RECIEN NACIDOS CON ALTA HOSPITALARIA - {str(fechaF).split('-')[0]}"
		sheet['A1'].alignment=Alignment(horizontal="center")
		sheet['A1'].border=self.borde_caja

		sheet.merge_cells('U1:V1')
		sheet['U1']="Desde: "+str(fechaI)+" -  Hasta: "+str(fechaF) 

		sheet.column_dimensions['A'].width = 8
		sheet.merge_cells('A2:A3')
		sheet['A2']="Nro"

		sheet.column_dimensions['B'].width = 8
		sheet.merge_cells('B2:B3')
		sheet['B2']="HCL"

		sheet.column_dimensions['C'].width = 30
		sheet.merge_cells('C2:C3')
		sheet['C2']="Nombres y Apellidos RN"


		sheet.column_dimensions['D'].width = 15
		sheet.merge_cells('D2:D3')
		sheet['D2']="Fecha Nacimiento"

		sheet.column_dimensions['E'].width = 10
		sheet.merge_cells('E2:E3')
		sheet['E2']="DNI RN"	

		
		sheet.column_dimensions['F'].width = 12
		sheet.merge_cells('F2:F3')
		sheet['F2']="SEXO"

		sheet.column_dimensions['G'].width = 8
		sheet.merge_cells('G2:G3')
		sheet['G2']="PESO RN"

		sheet.column_dimensions['H'].width = 8
		sheet.merge_cells('H2:H3')
		sheet['H2']="TALLA RN"

		sheet.column_dimensions['I'].width = 10
		sheet.merge_cells('I2:I3')
		sheet['I2']="EDAD GESTACIONAL"

		sheet.merge_cells('J2:J3')
		sheet['J2']="HB RN"

		sheet.merge_cells('K2:M2')
		sheet['K2']="CONDICION RN"

		sheet['K3']="RNAT"		
		sheet['L3']="RNPT"		
		sheet['M3']="BPN"

		sheet.merge_cells('N2:N3')	
		sheet['N2']="MELLIZOS"

		sheet.merge_cells('O2:O3')	
		sheet['O2']="FECHA TAMIZAJE NEONATAL"		

		sheet.merge_cells('P2:P3')	
		sheet['P2']="TIPO DE SEGURO"

		sheet.column_dimensions['Q'].width = 30
		sheet.merge_cells('Q2:Q3')	
		sheet['Q2']="NOMBRE DE LA MADRE"

		sheet.column_dimensions['R'].width = 20
		sheet.merge_cells('R2:R3')	
		sheet['R2']="PROCEDENCIA"

		sheet.column_dimensions['S'].width = 15
		sheet.merge_cells('S2:S3')	
		sheet['S2']="NRO CELULAR"

		sheet.column_dimensions['T'].width = 9
		sheet.merge_cells('T2:T3')	
		sheet['T2']="TIPO DE PARTO"

		sheet.merge_cells('U2:U3')	
		sheet['U2']="LUGAR DE NACIMIENTO"	


		n=4
		nro=1
		for valor in rows:

			HCL=valor.HCL						
			rows_Madre=obj_ConsultaGalen.query_Paciente(valor.DNI)					
			sheet['A'+str(n)]=nro
			sheet['B'+str(n)]=HCL
			rowsGalenNino=obj_ConsultaGalen.query_PacienteXHCL(HCL)
			if rowsGalenNino:
				sheet['C'+str(n)]=rowsGalenNino[0].PrimerNombre+" "+rowsGalenNino[0].ApellidoPaterno+" "+rowsGalenNino[0].ApellidoMaterno
				sheet['F'+str(n)]=rowsGalenNino[0].Descripcion

			peso=valor.PESO
			sheet['D'+str(n)]=valor.Fecha_Nacimiento
			sheet['E'+str(n)]=valor.CNV			
			sheet['G'+str(n)]=peso
			sheet['H'+str(n)]=valor.TALLA
			edadGesta=valor.EGESTACIONAL
			sheet['I'+str(n)]=edadGesta
			sheet['P'+str(n)]=valor.FINANCIAMIENTO

			if peso<2500:
				sheet['M'+str(n)]="SI"
			else:
				sheet['K'+str(n)]="SI"

			if not edadGesta==None:
				if int(edadGesta)<37:					
					sheet['L'+str(n)]="SI"
				else:
					
					sheet['L'+str(n)]="NO"

			idair=valor.Id_AIR
			rowsAlojamiento=obj_consulta.consulta_Tabla1("ALOJAMIENTO","Id_AIR",idair)
			if rowsAlojamiento:
				sheet['J'+str(n)]=rowsAlojamiento[0].HEMOGLOBINA
				sheet['O'+str(n)]=rowsAlojamiento[0].FECHA_TAMIZAJE
			if rows_Madre:
				sheet['Q'+str(n)]=rows_Madre[0].PrimerNombre+" "+rows_Madre[0].ApellidoPaterno+" "+rows_Madre[0].ApellidoMaterno
				sheet['R'+str(n)]=rows_Madre[0].Nombre
				sheet['S'+str(n)]=rows_Madre[0].Telefono
			sheet['T'+str(n)]=valor.tipo_Parto			
			
			nro=nro+1
			n=n+1

		wb.save(f"{address.name}")