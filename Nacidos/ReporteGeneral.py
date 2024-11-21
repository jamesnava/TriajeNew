from openpyxl import Workbook
from openpyxl.styles import Border,Side
from openpyxl.styles.alignment import Alignment
from Nacidos.consultaN import Consulta
from Consulta_Galen import queryGalen

class RGeneral(object):

	def __init__(self):
		self.obj_consultaN=Consulta()
		self.obj_consultaGalen=queryGalen()

	def General(self,event,direccion,fechaI,fechaF):
		
		rows=self.obj_consultaN.consulta_General(fechaI,fechaF)
		wb=Workbook()
		sheet=wb.active
		sheet.merge_cells('A1:K1')		
		sheet['A1']=f"REGISTRO DE RECIEN NACIDOS"
		sheet['A1'].alignment=Alignment(horizontal="center")
		

		sheet.merge_cells('L1:V1')
		sheet['L1']="Desde: "+str(fechaI)+" -  Hasta: "+str(fechaF) 

		sheet.merge_cells('A2:D2')
		sheet['A2']="DATOS DE LA MADRE"
		sheet['A2'].alignment=Alignment(horizontal="center")
		
		sheet.merge_cells('E2:W2')
		sheet['E2']="DATOS DEL RECIEN NACIDO"
		sheet['E2'].alignment=Alignment(horizontal="center")
		

		sheet['A3']="DNI MADRE"
		sheet['B3']="NOMBRES Y APELLIDOS"
		sheet['C3']="GRUPO_FACTOR"
		sheet['D3']="EDAD GEST"							
		sheet['E3']="HCL RN"		
		sheet['F3']="NOMBRES Y APELLIDOS"		
		sheet['G3']="CNV"		
		sheet['H3']="PESO"
		sheet['I3']="TALLA"
		sheet['J3']="PC"		
		sheet['K3']="PT"		
		sheet['L3']="PA"		
		sheet['M3']="PB"		
		sheet['N3']="EXA FI"		
		sheet['O3']="FUR"
		sheet['P3']="APGAR1"
		sheet['Q3']="APGAR5"
		sheet['R3']="APGAR10"		
		sheet['S3']="ASFIXIA"		
		sheet['T3']="GRUPO FACTOR"					
		sheet['U3']="Fecha Nacimiento"			
		sheet['V3']="REGISTRADO POR"	
		
		nro=4
		for valor in rows:
			dnipaciente=valor.DNI					
			sheet['A'+str(nro)]=dnipaciente
			#datos paciente
			rowspaciente=self.obj_consultaGalen.query_DatosPaciente(dnipaciente)
			paciente=rowspaciente[0].PrimerNombre+" "+rowspaciente[0].ApellidoPaterno+" "+rowspaciente[0].ApellidoMaterno
			sheet['B'+str(nro)]=paciente
			sheet['C'+str(nro)]=valor.GRUPO_FACTOR
			
			sheet['D'+str(nro)]=valor.EGESTACIONAL			
			hcl=valor.HCL
			
			rowsRN=self.obj_consultaGalen.query_PacienteXHCL(hcl)
			rndatos=rowsRN[0].PrimerNombre+" "+rowsRN[0].ApellidoPaterno+" "+rowsRN[0].ApellidoMaterno
			sheet['E'+str(nro)]=hcl
			sheet['F'+str(nro)]=rndatos
			sheet['G'+str(nro)]=valor.CNV
			sheet['H'+str(nro)]=valor.PESO
			sheet['I'+str(nro)]=valor.TALLA
			sheet['J'+str(nro)]=valor.PC
			sheet['K'+str(nro)]=valor.PT
			sheet['L'+str(nro)]=valor.PA
			sheet['M'+str(nro)]=valor.PB
			sheet['N'+str(nro)]=valor.EX_FI
			sheet['O'+str(nro)]=valor.FUR
			sheet['P'+str(nro)]=valor.APGAR_1
			sheet['Q'+str(nro)]=valor.APGAR_5
			sheet['R'+str(nro)]=valor.APGAR_10 if valor.APGAR_10!=-1 else ' '
			sheet['S'+str(nro)]=valor.ASFIXIA
			sheet['T'+str(nro)]=valor.GRUPORNA
			sheet['U'+str(nro)]=valor.Fecha_Nacimiento
			sheet['V'+str(nro)]=valor.RESPONSABLEATENCION			
			nro=nro+1			

		wb.save(f"{direccion}")

	def Interconsulta(self,event,direccion,fechaI,fechaF):
		
		rows=self.obj_consultaN.Interconsulta(fechaI,fechaF)
		wb=Workbook()
		sheet=wb.active
		sheet.merge_cells('A1:K1')		
		sheet['A1']=f"REPORTE DE INTERCONSULTAS"
		sheet['A1'].alignment=Alignment(horizontal="center")
		

		sheet.merge_cells('L1:V1')
		sheet['L1']="Desde: "+str(fechaI)+" -  Hasta: "+str(fechaF) 

		sheet.merge_cells('A2:C2')
		sheet['A2']="DATOS DE LA MADRE"
		sheet['A2'].alignment=Alignment(horizontal="center")
		
		sheet.merge_cells('D2:M2')
		sheet['D2']="DATOS DEL RECIEN NACIDO"
		sheet['D2'].alignment=Alignment(horizontal="center")		

		sheet['A3']="DNI MADRE"
		sheet['B3']="NOMBRES Y APELLIDOS"		
		sheet['C3']="EDAD GEST"							
		sheet['D3']="HCL RN"		
		sheet['E3']="NOMBRES Y APELLIDOS"		
		sheet['F3']="CNV"		
		sheet['G3']="PESO"
		sheet['H3']="TALLA"		
		sheet['I3']="APGAR1"
		sheet['J3']="APGAR5"
		sheet['K3']="APGAR10"						
		sheet['L3']="Fecha Nacimiento"
		sheet['M3']="Medico Responsable"			
		sheet['N3']="REGISTRADO POR"	
		
		nro=4
		for valor in rows:
			dnipaciente=valor.DNI					
			sheet['A'+str(nro)]=dnipaciente
			#datos paciente
			rowspaciente=self.obj_consultaGalen.query_DatosPaciente(dnipaciente)
			paciente=rowspaciente[0].PrimerNombre+" "+rowspaciente[0].ApellidoPaterno+" "+rowspaciente[0].ApellidoMaterno
			sheet['B'+str(nro)]=paciente			
			sheet['C'+str(nro)]=valor.EGESTACIONAL

			hcl=valor.HCL
			rowsRN=self.obj_consultaGalen.query_PacienteXHCL(hcl)
			rndatos=rowsRN[0].PrimerNombre+" "+rowsRN[0].ApellidoPaterno+" "+rowsRN[0].ApellidoMaterno
			sheet['D'+str(nro)]=hcl
			sheet['E'+str(nro)]=rndatos
			sheet['F'+str(nro)]=valor.CNV
			sheet['G'+str(nro)]=valor.PESO
			sheet['H'+str(nro)]=valor.TALLA			
			sheet['I'+str(nro)]=valor.APGAR_1
			sheet['J'+str(nro)]=valor.APGAR_5
			sheet['K'+str(nro)]=valor.APGAR_10 if valor.APGAR_10!=-1 else ' '			
			sheet['L'+str(nro)]=valor.Fecha_Nacimiento

			rowsmedico=self.obj_consultaGalen.query_EmpleadoDNI(valor.RESP_MEDICO_INTERCONSULTA)
			if len(rowsmedico):
				datosmedico=rowsmedico[0].Nombres+" "+rowsmedico[0].ApellidoPaterno+" "+rowsmedico[0].ApellidoMaterno				
				sheet['M'+str(nro)]=datosmedico
			sheet['N'+str(nro)]=valor.RESPONSABLEATENCION			
			nro=nro+1			

		wb.save(f"{direccion}")

