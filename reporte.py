from openpyxl import Workbook
from openpyxl.styles import Border,Side
from openpyxl.styles.alignment import Alignment
from xml.etree.ElementTree import Element, SubElement, tostring
from et_xmlfile import xmlfile
import Consulta_Triaje,Consulta_Galen
from tkinter import messagebox

class Reporte(object):
	def __init__(self):
		double_border_side=Side(border_style="thin")
		self.borde_caja=Border(top=double_border_side,right=double_border_side,bottom=double_border_side,left=double_border_side)
		self.borde_superior=Border(top=Side(border_style='thick'))
		self.obj_consulta=Consulta_Triaje.queryTriaje()
		self.obj_consultaG=Consulta_Galen.queryGalen()
	def Genera_RDatos(self,codigo,address,fecha):
		aux=False
		rows=self.obj_consulta.datos_HojaV2(codigo,fecha)		

		if len(rows)>0:
			self.Report_DigitD(rows,address,fecha,fecha,codigo)
			aux=True
		else:
			aux=False
		return aux

	def Genera_RDatosGeneral(self,dni,address,fechaI,fechaF,Especialidad):
		aux=False
		rows=self.obj_consulta.datos_HojaXDias(dni,fechaI,fechaF,Especialidad)

		if len(rows)>0:
			self.Report_DigitD(rows,address,fechaI,fechaF,dni)
			aux=True
		else:
			aux=False
		return aux

	def Report_DigitD(self,rows,address,fechaI,fechaF,dni):
		#try:			
		wb=Workbook()
		sheet=wb.active
		#Encabezado de la hoja HIS
		sheet.merge_cells('A1:Q1')
		#sheet['A1'].border=self.borde_caja
		sheet['A1']="REGISTRO DIARIO DE ANTENCION Y OTRAS ACTIVIDADES DE SALUD"
		sheet['A1'].alignment=Alignment(horizontal="center")
		sheet['A3']="AÑO"
		sheet['A3'].border=self.borde_caja
		sheet['B3']="MES"
		sheet['B3'].border=self.borde_caja
		sheet.merge_cells('C3:F3')
		sheet['C3']="NOMBRE DEL ESTABLECIMIENTO"
		sheet['C3'].border=self.borde_caja
		sheet['D3'].border=self.borde_caja
		sheet['E3'].border=self.borde_caja
		sheet['F3'].border=self.borde_caja
		sheet.merge_cells('G3:K3')
		sheet['G3']="UNIDAD DE PRODUCTORA DE SERVICIO"
		sheet['G3'].border=self.borde_caja
		sheet.merge_cells('L3:M3')
		sheet['L3']="DNI RESP"
		sheet['L3'].border=self.borde_caja
		sheet.merge_cells('N3:O3')
		sheet['N3']="RESPONSABLE"
		sheet.merge_cells('P3:Q3')
		sheet['P3']="DESDE - HASTA"
		sheet['H3'].border=self.borde_caja
		sheet['I3'].border=self.borde_caja
		sheet['J3'].border=self.borde_caja
		sheet['K3'].border=self.borde_caja
		sheet['L3'].border=self.borde_caja
		sheet['N3'].border=self.borde_caja
		sheet['M3'].border=self.borde_caja
		sheet['O3'].border=self.borde_caja
		sheet['P3'].border=self.borde_caja
		sheet['Q3'].border=self.borde_caja
		#combinar celdas
		sheet.merge_cells('C4:F4')			
		sheet.merge_cells('G4:K4')
		sheet.merge_cells('L4:M4')
		sheet.merge_cells('N4:O4')
		sheet.merge_cells('P4:Q4')
		sheet['A6'].border=self.borde_caja		
		sheet['A6']="DNI"
		sheet['B6'].border=self.borde_caja
		sheet['B6']="HISTORIA"
		sheet['C6'].border=self.borde_caja
		sheet.column_dimensions['C'].width = 25
		sheet['C6']="PACIENTE"
		sheet['D6'].border=self.borde_caja
		sheet['D6']="FINANC."
		sheet['E6'].border=self.borde_caja
		sheet.column_dimensions['E'].width = 18
		sheet['E6']="DISTRITO PROC."
		sheet['F6'].border=self.borde_caja
		sheet['F6']="CENTRO POBL."
		sheet['G6'].border=self.borde_caja
		sheet.column_dimensions['G'].width = 5
		sheet['G6']="PC"
		sheet['H6'].border=self.borde_caja
		sheet.column_dimensions['H'].width = 5
		sheet['H6']="PAB"
		sheet['I6'].border=self.borde_caja
		sheet['I6']="PESO"
		sheet['J6'].border=self.borde_caja
		sheet['J6']="TALLA"
		sheet['K6'].border=self.borde_caja
		sheet['K6']="HB"
		sheet['L6'].border=self.borde_caja
		sheet['L6']="ESTABLECIMIENTO"
		sheet['M6'].border=self.borde_caja
		sheet['M6']="SERVICIO"
		sheet.column_dimensions['N'].width = 30
		sheet['N6'].border=self.borde_caja
		sheet['N6']="DIAGNOSTICO"
		sheet['O6'].border=self.borde_caja
		sheet['O6']="TIPO DX"
		sheet['P6'].border=self.borde_caja				
		sheet['P6']="LAB"
		sheet['Q6'].border=self.borde_caja
		sheet['Q6']="CIE"		
		sheet['C4']="HOSPITAL SUB REGIONAL DE ANDAHUAYLAS"
		rowsG=self.obj_consultaG.query_MedicoData(fechaI,fechaF,dni)
		if not rowsG:
			sheet['G4']=""
			rowsG=self.obj_consultaG.query_EmpleadoDNI(dni)
			sheet['N4']=rowsG[0].Nombres+" "+rowsG[0].ApellidoPaterno+" "+rowsG[0].ApellidoMaterno
		else:
			sheet['G4']=rowsG[0].Nombre
			sheet['N4']=rowsG[0].Nombres+" "+rowsG[0].map+" "+rowsG[0].mam
		
		
		sheet['P4']=str(fechaI)+"=>"+str(fechaF)
		nro=7
		for datos in rows:
			#agregando a la cabecera
			sheet['A4']=datos.FechaIngreso[:4]			
			sheet['B4']=datos.FechaIngreso[5:7]
			#sheet['C4']=datos.ESTABLECIMIENTO
			sheet['L4']=datos.DNI_USER			
			sheet['A'+str(nro)]=datos.DNI_PAC

			rowsDatos=self.obj_consultaG.query_datosPaciente(datos.DNI_PAC)
			if rowsDatos:
				sheet['B'+str(nro)]=rowsDatos[0].NroHistoriaClinica
				sheet['C'+str(nro)]=rowsDatos[0].PrimerNombre+" "+rowsDatos[0].ApellidoPaterno+" "+rowsDatos[0].ApellidoMaterno				
				sheet['E'+str(nro)]=rowsDatos[0].Nombre
			#dias
			sheet['A'+str(nro+1)]=str(datos.FechaIngreso)[5:]

			#sheet['D'+str(nro)]=datos.FINACIAMIENTO
			sheet['G'+str(nro)]=datos.PC
			sheet['H'+str(nro)]=datos.PAB
			sheet['I'+str(nro)]=datos.Peso
			sheet['J'+str(nro)]=datos.Talla
			sheet['K'+str(nro)]=datos.Hb
			sheet['L'+str(nro)]=datos.Establecimiento
			sheet['M'+str(nro)]=datos.Servicio
			rows_deta=self.obj_consulta.query_DIAGNOSTICOS(datos.ID_DETA)
			numeracion=0
			sheet['A'+str(nro)].border=self.borde_superior
			sheet['B'+str(nro)].border=self.borde_superior
			sheet['C'+str(nro)].border=self.borde_superior
			sheet['D'+str(nro)].border=self.borde_superior
			sheet['E'+str(nro)].border=self.borde_superior
			sheet['F'+str(nro)].border=self.borde_superior
			sheet['G'+str(nro)].border=self.borde_superior
			sheet['H'+str(nro)].border=self.borde_superior
			sheet['I'+str(nro)].border=self.borde_superior
			sheet['J'+str(nro)].border=self.borde_superior
			sheet['K'+str(nro)].border=self.borde_superior
			sheet['L'+str(nro)].border=self.borde_superior
			sheet['M'+str(nro)].border=self.borde_superior
			sheet['N'+str(nro)].border=self.borde_superior
			sheet['O'+str(nro)].border=self.borde_superior
			sheet['P'+str(nro)].border=self.borde_superior
			sheet['Q'+str(nro)].border=self.borde_superior
			for dx in rows_deta:
				sheet['N'+str(nro+numeracion)].border=self.borde_caja
				sheet['N'+str(nro+numeracion)]=dx.Descripcion
				sheet['O'+str(nro+numeracion)].border=self.borde_caja
				sheet['O'+str(nro+numeracion)]=dx.TipDx
				sheet['P'+str(nro+numeracion)].border=self.borde_caja
				sheet['P'+str(nro+numeracion)]=dx.Lab
				sheet['Q'+str(nro+numeracion)].border=self.borde_caja
				sheet['Q'+str(nro+numeracion)]=dx.CODCIE
				numeracion=numeracion+1
			nro=nro+numeracion+1			
		wb.save(f"{address.name}")
		#except Exception as e:
		#	messagebox.showerror("Alerta",f"{e}")

